# custom_api/custom_api/api/inventory_import.py
#
# Inventory Import API for Frappe / ERPNext
# Supports CSV and XLSX. Creates Items (if missing) + Opening Stock via
# Stock Reconciliation (grouped by warehouse).
#
# Expected columns in file (header row required):
#   item_code, item_name, item_group, uom, warehouse, opening_qty
# Optional columns:
#   stock_uom, valuation_rate, description, brand
#   batch_no      -> only used if item has_batch_no=1 (auto-creates a Batch if not given)
#   serial_nos    -> only used if item has_serial_no=1 (comma-separated, count must equal opening_qty)
#
# Usage:
#   1) REST: POST /api/method/custom_api.custom_api.api.inventory_import.import_inventory
#      - multipart/form-data with key "file"  (direct upload), OR
#      - JSON body { "file_url": "/private/files/xyz.csv" } (file already uploaded via /api/method/upload_file)
#   2) Desk UI: see inventory_import.js (Workspace page with upload button)

import frappe
from frappe import _
from collections import defaultdict
import csv
from openpyxl import load_workbook

REQUIRED_COLUMNS = ["item_code", "item_name", "item_group", "uom", "warehouse", "opening_qty"]


# ---------------------------------------------------------------------
# File parsing
# ---------------------------------------------------------------------
def parse_file(file_doc):
    """file_doc: Frappe File document. Returns list[dict] rows."""
    file_path = file_doc.get_full_path()
    ext = file_doc.file_name.rsplit(".", 1)[-1].lower()

    rows = []
    if ext == "csv":
        with open(file_path, "r", encoding="utf-8-sig") as f:
            reader = csv.DictReader(f)
            for row in reader:
                rows.append({(k or "").strip(): (v or "").strip() for k, v in row.items()})
    elif ext in ("xlsx", "xls"):
        wb = load_workbook(file_path, data_only=True)
        ws = wb.active
        headers = [str(c.value).strip() if c.value is not None else "" for c in ws[1]]
        for r in ws.iter_rows(min_row=2, values_only=True):
            if r is None or all(v is None for v in r):
                continue
            row_dict = {}
            for h, v in zip(headers, r):
                row_dict[h] = "" if v is None else str(v).strip()
            if any(row_dict.values()):
                rows.append(row_dict)
    else:
        frappe.throw(_("Unsupported file type: {0}. Use CSV or XLSX.").format(ext))

    return rows


# ---------------------------------------------------------------------
# Validation
# ---------------------------------------------------------------------
def validate_row(row, row_num, errors):
    missing = [c for c in REQUIRED_COLUMNS if not row.get(c)]
    if missing:
        errors.append(f"Row {row_num}: Missing {', '.join(missing)}")
        return False

    try:
        float(row.get("opening_qty") or 0)
    except ValueError:
        errors.append(f"Row {row_num}: Invalid opening_qty '{row.get('opening_qty')}'")
        return False

    if row.get("valuation_rate"):
        try:
            float(row["valuation_rate"])
        except ValueError:
            errors.append(f"Row {row_num}: Invalid valuation_rate '{row.get('valuation_rate')}'")
            return False

    return True


# ---------------------------------------------------------------------
# Item creation
# ---------------------------------------------------------------------
def get_or_create_item(row):
    item_code = row["item_code"]
    if frappe.db.exists("Item", item_code):
        return item_code

    if not frappe.db.exists("Item Group", row["item_group"]):
        frappe.throw(_("Item Group '{0}' does not exist (item {1})").format(row["item_group"], item_code))

    uom = row.get("stock_uom") or row["uom"]
    if not frappe.db.exists("UOM", uom):
        frappe.throw(_("UOM '{0}' does not exist (item {1})").format(uom, item_code))

    item = frappe.new_doc("Item")
    item.item_code = item_code
    item.item_name = row.get("item_name") or item_code
    item.item_group = row["item_group"]
    item.stock_uom = uom
    item.is_stock_item = 1
    item.description = row.get("description") or row.get("item_name") or item_code
    if row.get("brand"):
        item.brand = row["brand"]
    item.insert(ignore_permissions=True)
    return item.name


# ---------------------------------------------------------------------
# Serial / Batch handling (required by newer ERPNext for tracked items)
# ---------------------------------------------------------------------
def get_serial_batch_bundle(item_code, warehouse, qty, batch_no=None, serial_nos=None):
    """
    Returns a Serial and Batch Bundle name if the item is serial/batch tracked,
    else None (plain qty-based reconciliation).
    """
    item_meta = frappe.get_cached_doc("Item", item_code)
    if not (item_meta.has_batch_no or item_meta.has_serial_no):
        return None

    bundle = frappe.new_doc("Serial and Batch Bundle")
    bundle.item_code = item_code
    bundle.warehouse = warehouse
    bundle.voucher_type = "Stock Reconciliation"
    bundle.type_of_transaction = "Inward"

    if item_meta.has_serial_no:
        if not serial_nos:
            frappe.throw(
                _("Item {0} is Serial No tracked. Provide a 'serial_nos' column (comma-separated serials, count must match opening_qty).").format(item_code)
            )
        serials = [s.strip() for s in serial_nos.split(",") if s.strip()]
        if len(serials) != int(qty):
            frappe.throw(
                _("Item {0}: {1} serial numbers given but opening_qty is {2}").format(item_code, len(serials), int(qty))
            )
        for sn in serials:
            bundle.append("entries", {"serial_no": sn, "qty": 1, "warehouse": warehouse})
    else:
        # Batch tracked (no serials)
        b_no = batch_no
        if not b_no or not frappe.db.exists("Batch", {"item": item_code, "name": b_no}):
            batch = frappe.new_doc("Batch")
            batch.item = item_code
            if batch_no:
                batch.batch_id = batch_no
            batch.insert(ignore_permissions=True)
            b_no = batch.name
        bundle.append("entries", {"batch_no": b_no, "qty": qty, "warehouse": warehouse})

    bundle.insert(ignore_permissions=True)
    return bundle.name


# ---------------------------------------------------------------------
# Opening stock via Stock Reconciliation (grouped per warehouse)
# ---------------------------------------------------------------------
def create_stock_reconciliations(valid_rows):
    by_warehouse = defaultdict(list)
    for row in valid_rows:
        by_warehouse[row["warehouse"]].append(row)

    default_company = frappe.defaults.get_user_default("Company") or frappe.db.get_single_value(
        "Global Defaults", "default_company"
    )

    created = []
    for warehouse, rows in by_warehouse.items():
        if not frappe.db.exists("Warehouse", warehouse):
            frappe.throw(_("Warehouse '{0}' does not exist").format(warehouse))

        sr = frappe.new_doc("Stock Reconciliation")
        sr.purpose = "Opening Stock"
        sr.company = default_company

        for row in rows:
            qty = float(row["opening_qty"])
            item_row = {
                "item_code": row["item_code_final"],
                "warehouse": warehouse,
                "qty": qty,
            }
            if row.get("valuation_rate"):
                item_row["valuation_rate"] = float(row["valuation_rate"])

            bundle_name = get_serial_batch_bundle(
                row["item_code_final"],
                warehouse,
                qty,
                batch_no=row.get("batch_no"),
                serial_nos=row.get("serial_nos"),
            )
            if bundle_name:
                item_row["serial_and_batch_bundle"] = bundle_name
                item_row["use_serial_batch_fields"] = 0

            sr.append("items", item_row)

        sr.insert(ignore_permissions=True)
        try:
            sr.submit()
            created.append({"warehouse": warehouse, "name": sr.name, "status": "submitted"})
        except Exception:
            frappe.log_error(frappe.get_traceback(), f"Stock Reconciliation submit failed - {warehouse}")
            created.append({"warehouse": warehouse, "name": sr.name, "status": "draft (submit failed - check Error Log)"})

    return created


# ---------------------------------------------------------------------
# Whitelisted entry point
# ---------------------------------------------------------------------
@frappe.whitelist()
def import_inventory(file_url=None):
    """
    Import Items + Opening Stock from an uploaded CSV/XLSX file.
    Call with file_url of an already-uploaded File doc, OR attach the file
    directly (multipart/form-data, field name "file").
    """
    if not any(role in frappe.get_roles() for role in ("System Manager", "Stock Manager", "Item Manager")):
        frappe.throw(_("Not permitted"), frappe.PermissionError)

    file_doc = None
    if file_url:
        file_doc = frappe.get_doc("File", {"file_url": file_url})
    elif frappe.request and frappe.request.files:
        upload = frappe.request.files.get("file")
        if not upload:
            frappe.throw(_("No file uploaded"))
        file_doc = frappe.get_doc({
            "doctype": "File",
            "file_name": upload.filename,
            "content": upload.stream.read(),
            "is_private": 1,
        })
        file_doc.save(ignore_permissions=True)
    else:
        frappe.throw(_("Provide file_url or upload a file"))

    rows = parse_file(file_doc)
    if not rows:
        frappe.throw(_("File has no data rows"))

    # Large file -> hand off to background job, return job status immediately
    if len(rows) > 500:
        frappe.enqueue(
            "custom_api.custom_api.api.inventory_import._run_import",
            queue="long",
            rows=rows,
            user=frappe.session.user,
        )
        return {
            "success": True,
            "queued": True,
            "message": f"{len(rows)} rows queued for background import. Check Error Log / Stock Reconciliation list for results.",
        }

    return _run_import(rows)


def _run_import(rows, user=None):
    if user:
        frappe.set_user(user)

    errors, valid_rows = [], []
    for i, row in enumerate(rows, start=2):
        if validate_row(row, i, errors):
            valid_rows.append(row)

    if errors and not valid_rows:
        return {"success": False, "errors": errors, "reconciliations": []}

    created_items, good_rows = [], []
    for row in valid_rows:
        try:
            item_name_final = get_or_create_item(row)
            row["item_code_final"] = item_name_final
            good_rows.append(row)
            created_items.append(item_name_final)
        except Exception as e:
            frappe.db.rollback()
            errors.append(f"Item '{row.get('item_code')}': {e}")

    reconciliations = []
    if good_rows:
        frappe.db.commit()
        reconciliations = create_stock_reconciliations(good_rows)

    return {
        "success": True,
        "total_rows": len(rows),
        "items_processed": len(created_items),
        "unique_items": len(set(created_items)),
        "reconciliations": reconciliations,
        "errors": errors,
    }